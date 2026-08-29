"""Integration tests for dcmspec_explorer.services.service_mediator.

These tests drive the real IODListLoaderWorker/IODModelLoaderWorker through the
mediator's real start_worker, i.e. a real threading.Thread and a real QTimer polling
loop, with a FakeModel whose load methods return or raise immediately so the
background thread finishes fast and the tests stay deterministic and non-flaky.
"""

import pathlib

from anytree import Node
from openpyxl import load_workbook

from dcmspec.progress import Progress
from dcmspec.spec_model import SpecModel

from dcmspec_explorer.services.service_mediator import (
    IODListLoaderServiceMediator,
    IODModelLoaderServiceMediator,
    IODExportServiceMediator,
)


class FakeModel:
    """Fake model whose load methods return/raise immediately, optionally reporting progress first."""

    def __init__(self, iod_list=None, iod_model=None, error=None, progress=None):
        """Initialize with the value or exception each load method should produce."""
        self.iod_list = iod_list
        self.iod_model = iod_model
        self.error = error
        self.progress = progress

    def load_iod_list(self, force_download, progress_observer):
        """Report progress if configured, then return iod_list or raise the canned error."""
        return self._result(progress_observer, self.iod_list)

    def load_iod_model(self, table_id, logger, progress_observer):
        """Report progress if configured, then return iod_model or raise the canned error."""
        return self._result(progress_observer, self.iod_model)

    def _result(self, progress_observer, value):
        """Report progress if configured, then raise the canned error or return value."""
        if self.progress is not None:
            progress_observer(self.progress)
        if self.error is not None:
            raise self.error
        return value


class TestIODListLoaderServiceMediatorHappyPath:
    """Tests for IODListLoaderServiceMediator's signal emission and cleanup on real worker events."""

    def test_progress_event_emits_progress_signal_with_payload(self, qtbot, fake_logger):
        """A progress event reported by the model emits iodlist_progress_signal with that Progress."""
        progress = Progress(percent=50)
        model = FakeModel(iod_list=["entry"], progress=progress)
        mediator = IODListLoaderServiceMediator(model=model, logger=fake_logger)

        with qtbot.waitSignal(mediator.iodlist_progress_signal, timeout=1000) as blocker:
            mediator.start_iodlist_worker()

        emitted_mediator, emitted_progress = blocker.args
        assert emitted_mediator is mediator
        assert emitted_progress is progress

    def test_loaded_event_emits_loaded_signal_and_cleans_up(self, qtbot, fake_logger):
        """A successful load emits iodlist_loaded_signal, stops the poll timer, and drops worker/thread."""
        model = FakeModel(iod_list=["entry1", "entry2"])
        mediator = IODListLoaderServiceMediator(model=model, logger=fake_logger)

        with qtbot.waitSignal(mediator.iodlist_loaded_signal, timeout=1000) as blocker:
            mediator.start_iodlist_worker()

        emitted_mediator, entries = blocker.args
        assert emitted_mediator is mediator
        assert entries == ["entry1", "entry2"]
        assert not hasattr(mediator, "_worker")
        assert not hasattr(mediator, "_thread")
        assert not mediator._poll_timer.isActive()

    def test_error_event_emits_error_signal_and_cleans_up(self, qtbot, fake_logger):
        """A raised error emits iodlist_error_signal with its message and does the same cleanup."""
        model = FakeModel(error=RuntimeError("boom"))
        mediator = IODListLoaderServiceMediator(model=model, logger=fake_logger)

        with qtbot.waitSignal(mediator.iodlist_error_signal, timeout=1000) as blocker:
            mediator.start_iodlist_worker()

        emitted_mediator, message = blocker.args
        assert emitted_mediator is mediator
        assert message == "boom"
        assert not hasattr(mediator, "_worker")
        assert not hasattr(mediator, "_thread")
        assert not mediator._poll_timer.isActive()


class TestIODModelLoaderServiceMediatorHappyPath:
    """Tests for IODModelLoaderServiceMediator's signal emission and cleanup on real worker events."""

    def test_progress_event_emits_progress_signal_with_payload(self, qtbot, fake_logger):
        """A progress event reported by the model emits iodmodel_progress_signal with that Progress."""
        progress = Progress(percent=50)
        model = FakeModel(iod_model="some_spec_model", progress=progress)
        mediator = IODModelLoaderServiceMediator(model=model, logger=fake_logger)

        with qtbot.waitSignal(mediator.iodmodel_progress_signal, timeout=1000) as blocker:
            mediator.start_iodmodel_worker(table_id="table_A.2-1")

        emitted_mediator, emitted_progress = blocker.args
        assert emitted_mediator is mediator
        assert emitted_progress is progress

    def test_loaded_event_emits_loaded_signal_and_cleans_up(self, qtbot, fake_logger):
        """A successful load emits iodmodel_loaded_signal, stops the poll timer, and drops worker/thread."""
        model = FakeModel(iod_model="some_spec_model")
        mediator = IODModelLoaderServiceMediator(model=model, logger=fake_logger)

        with qtbot.waitSignal(mediator.iodmodel_loaded_signal, timeout=1000) as blocker:
            mediator.start_iodmodel_worker(table_id="table_A.2-1")

        emitted_mediator, iod_model = blocker.args
        assert emitted_mediator is mediator
        assert iod_model == "some_spec_model"
        assert not hasattr(mediator, "_worker")
        assert not hasattr(mediator, "_thread")
        assert not mediator._poll_timer.isActive()

    def test_error_event_emits_error_signal_and_cleans_up(self, qtbot, fake_logger):
        """A raised error emits iodmodel_error_signal with its message and does the same cleanup."""
        model = FakeModel(error=ValueError("bad table_id"))
        mediator = IODModelLoaderServiceMediator(model=model, logger=fake_logger)

        with qtbot.waitSignal(mediator.iodmodel_error_signal, timeout=1000) as blocker:
            mediator.start_iodmodel_worker(table_id="table_A.2-1")

        emitted_mediator, message = blocker.args
        assert emitted_mediator is mediator
        assert message == "bad table_id"
        assert not hasattr(mediator, "_worker")
        assert not hasattr(mediator, "_thread")
        assert not mediator._poll_timer.isActive()


class TestOverlappingWorkerStartsHazard:
    """Regression tests characterizing the lack of any single-worker-in-flight guard.

    Both `initialize_treeview` and `_on_reload_clicked` call `start_iodlist_worker`
    unconditionally, with no check for a worker already in flight on the same mediator.
    These tests pin down the two concrete consequences of that missing guard.
    """

    def test_starting_a_second_worker_before_first_completes_orphans_the_first_timer(self, qtbot, fake_logger):
        """A second start_worker call before the first is polled leaves the first QTimer orphaned and running."""
        mediator = IODListLoaderServiceMediator(model=FakeModel(iod_list=["entry"]), logger=fake_logger)

        mediator.start_iodlist_worker()
        first_timer = mediator._poll_timer

        mediator.start_iodlist_worker()

        assert mediator._poll_timer is not first_timer
        assert first_timer.isActive()
        first_timer.stop()  # avoid leaking a live repeating timer into later tests in this session

    def test_cleanup_worker_thread_swallows_error_on_double_cleanup(self, qtbot, fake_logger):
        """Calling cleanup_worker_thread twice in a row does not raise, thanks to its broad except Exception."""
        mediator = IODListLoaderServiceMediator(model=FakeModel(), logger=fake_logger)

        mediator.cleanup_worker_thread()
        mediator.cleanup_worker_thread()


class TestIODExportServiceMediatorHappyPath:
    """Tests for IODExportServiceMediator's signal emission and cleanup on real worker events.

    Unlike the loader mediators above, these drive the real IODExportWorker against a real
    (minimal) SpecModel and the real dcmspec IODSpecPrinter, so the assertions cover the actual
    file written to disk, not just the emitted signal payload.
    """

    @staticmethod
    def _make_iod_model():
        """Return a minimal real SpecModel with one module and one attribute, for real IODSpecPrinter export."""
        metadata = Node("metadata")
        metadata.header = ["Attr1", "Attr2"]
        metadata.column_to_attr = {0: "attr1", 1: "attr2"}
        content = Node("content")
        module_node = Node("module1", parent=content)
        module_node.module = "Patient"
        module_node.usage = "M"
        attr_node = Node("attr", parent=module_node)
        attr_node.attr1 = "Value1"
        attr_node.attr2 = "Value2"
        model = SpecModel(metadata=metadata, content=content)
        model._is_include = lambda node: False
        model._is_title = lambda node: False
        return model

    def test_loaded_event_writes_real_csv_file_and_cleans_up(self, qtbot, fake_logger, tmp_path):
        """A real csv export via IODSpecPrinter writes the file and emits iodexport_loaded_signal."""
        output_path = str(tmp_path / "export.csv")
        mediator = IODExportServiceMediator(model=None, logger=fake_logger)

        with qtbot.waitSignal(mediator.iodexport_loaded_signal, timeout=1000) as blocker:
            mediator.start_export_worker(iod_model=self._make_iod_model(), fmt="csv", output_path=output_path)

        emitted_mediator, emitted_path = blocker.args
        assert emitted_mediator is mediator
        assert emitted_path == output_path
        content = pathlib.Path(output_path).read_text(encoding="utf-8")
        assert "Value1" in content
        assert "Value2" in content
        assert not hasattr(mediator, "_worker")
        assert not hasattr(mediator, "_thread")
        assert not mediator._poll_timer.isActive()

    @staticmethod
    def _make_iod_model_with_html_description():
        """Return a real SpecModel shaped like dcmspec-explorer's actual module attribute table.

        Uses the real column_to_attr ({0: "elem_name", 1: "elem_tag", 2: "elem_type",
        3: "elem_description"}) from Model.load_iod_model, with an HTML elem_description like the
        one PS3.3's XHTML actually produces, to exercise the real end-to-end HTML-to-text
        conversion path through IODExportWorker._to_plain_text_model.
        """
        metadata = Node("metadata")
        metadata.header = ["Name", "Tag", "Type", "Description"]
        metadata.column_to_attr = {0: "elem_name", 1: "elem_tag", 2: "elem_type", 3: "elem_description"}
        content = Node("content")
        module_node = Node("module1", parent=content)
        module_node.module = "Patient"
        module_node.usage = "M"
        attr_node = Node("attr", parent=module_node)
        attr_node.elem_name = "Patient's Name"
        attr_node.elem_tag = "(0010,0010)"
        attr_node.elem_type = "2"
        attr_node.elem_description = '<p>\n<a id="para_427a" shape="rect"/>Patient\'s full name.</p>'
        model = SpecModel(metadata=metadata, content=content)
        model._is_include = lambda node: False
        model._is_title = lambda node: False
        return model

    def test_loaded_event_converts_html_description_to_plain_text(self, qtbot, fake_logger, tmp_path):
        """The exported CSV contains the plain-text description, not the raw HTML markup."""
        output_path = str(tmp_path / "export_description.csv")
        mediator = IODExportServiceMediator(model=None, logger=fake_logger)

        with qtbot.waitSignal(mediator.iodexport_loaded_signal, timeout=1000):
            mediator.start_export_worker(
                iod_model=self._make_iod_model_with_html_description(), fmt="csv", output_path=output_path
            )

        content = pathlib.Path(output_path).read_text(encoding="utf-8")
        assert "Patient's full name." in content
        assert "<p>" not in content
        assert "<a id=" not in content

    def test_loaded_event_writes_real_xlsx_file(self, qtbot, fake_logger, tmp_path):
        """A real xlsx export via IODSpecPrinter writes a readable workbook and emits the loaded signal."""
        output_path = str(tmp_path / "export.xlsx")
        mediator = IODExportServiceMediator(model=None, logger=fake_logger)

        with qtbot.waitSignal(mediator.iodexport_loaded_signal, timeout=1000) as blocker:
            mediator.start_export_worker(iod_model=self._make_iod_model(), fmt="xlsx", output_path=output_path)

        _, emitted_path = blocker.args
        assert emitted_path == output_path
        workbook = load_workbook(output_path)
        sheet = workbook[workbook.sheetnames[0]]
        values = [cell.value for row in sheet.iter_rows() for cell in row]
        assert "Value1" in values

    def test_error_event_emits_error_signal_and_cleans_up(self, qtbot, fake_logger, tmp_path):
        """An unsupported format emits iodexport_error_signal with the failure message."""
        mediator = IODExportServiceMediator(model=None, logger=fake_logger)

        with qtbot.waitSignal(mediator.iodexport_error_signal, timeout=1000) as blocker:
            mediator.start_export_worker(
                iod_model=self._make_iod_model(), fmt="pdf", output_path=str(tmp_path / "export.pdf")
            )

        _, message = blocker.args
        assert message == "Unsupported export format: pdf"
        assert not hasattr(mediator, "_worker")
        assert not hasattr(mediator, "_thread")
        assert not mediator._poll_timer.isActive()
